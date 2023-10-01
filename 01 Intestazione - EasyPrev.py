import tkinter as tk
import openpyxl
from tkinter import messagebox
from tkinter import ttk, filedialog
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl import load_workbook


class ExcelEditor:
 
 
    def crea_menu(self):
        self.menu = tk.Menu(self.root)
        self.root.config(menu=self.menu)

        self.filemenu = tk.Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="File", menu=self.filemenu)
        self.filemenu.add_command(label="Carica File Excel", command=self.carica_file_excel)
        self.filemenu.add_command(label="Salva Excel", command=self.modifica_excel)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Informazioni", command=self.mostra_info)  # Nuovo comando per mostrare informazioni

    def mostra_info(self):
        messagebox.showinfo("EasyPrev AutoeXcel 1.0", " EasyPrev AutoeXcel è un software di preventivazione semi-automatizzata basata su python ed excel. Ultimo aggiornamnto 30.09.23.\n\n Programma realizzato per\n•Domenico Bertollini•\nSviluppato da\n•Luca Gualandri•")
        
    def __init__(self, root):
        self.root = root
        self.root.title("Intestazione - EasyPrev")
        self.root.geometry("700x500")
        self.foglio = None
        self.bordo_sottolineato = NamedStyle(name="bordo_sottolineato")
        self.bordo_sottolineato.font = Font(underline="single")
        bordo = Border(bottom=Side(style="thin"))
        self.bordo_sottolineato.border = bordo

        self.crea_interfaccia()
        self.crea_menu()

    def carica_file_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("File Excel", "*.xlsx")])
        if file_path:
            self.wb = load_workbook(file_path)  # Conserva l'intero workbook invece di solo il foglio attivo
            self.foglio = self.wb.active
            
    def ottieni_larghezze_colonne(self, foglio_originale):
        # Ritorna un dizionario con le larghezze delle colonne del foglio originale
        larghezze_colonne = {}
        for col_letter, col_dim in foglio_originale.column_dimensions.items():
            larghezze_colonne[col_letter] = col_dim.width
        return larghezze_colonne
    
    def imposta_larghezza_colonne(self, larghezze_colonne):
        # Imposta le larghezze delle colonne nel foglio corrente in base al dizionario fornito
        if self.foglio and larghezze_colonne:
            for col_letter, width in larghezze_colonne.items():
                self.foglio.column_dimensions[col_letter].width = width

    def crea_copia_excel(self):
        if self.foglio:
            for row in self.foglio.iter_rows(values_only=False):
                for cell in row:
                    if cell.font.underline:
                        bordo_sottolineato = NamedStyle(name="bordo_sottolineato")
                        bordo = Side(style='thin', color='000000')  # Stile sottolineato con colore nero
                        bordo_sottolineato.border = Border(bottom=bordo)  # Applica il bordo solo sulla parte inferiore

                        # Usa lo stile nella cella
                        cell.style = bordo_sottolineato

    def aggiungi_righe_excel(self, num_righe):
        if self.foglio:
            riga_16 = self.foglio[16]  # Riga 16 da cui copiare la formattazione
            for _ in range(num_righe):
                nuova_riga = self.foglio.insert_rows(17)  # Inserisci una nuova riga dopo la riga 16
                for index, (cella_sorgente, cella_destinazione) in enumerate(zip(riga_16, nuova_riga)):
                    if cella_sorgente.has_style:
                        cella_destinazione._style = copy(cella_sorgente._style)
                    col_letter = get_column_letter(index + 1)  # L'indice è 0-based, quindi aggiungi 1
                    self.foglio.column_dimensions[col_letter].width = \
                        self.foglio.column_dimensions[col_letter].width



    def compila_informazioni(self, luogo_data, informazione2, email, residenza, cap, attenzione, copia, oggetto):
        if self.foglio:
            self.foglio['A9'] = luogo_data
            self.foglio['A12'] = informazione2
            self.foglio['D11'] = email
            self.foglio['D12'] = residenza
            self.foglio['D13'] = cap
            self.foglio['D10'] = attenzione
            self.foglio['A13'] = copia

            # Suddividi il testo su più righe e aggiungi le righe necessarie
            if oggetto:
                righe_oggetto = ["           " + line for line in oggetto.split('\n')]
                for i, riga in enumerate(righe_oggetto, start=16):
                    if self.is_cell_merged(self.foglio, i, 2):
                        # Trova la cella in alto a sinistra dell'intervallo unito e imposta il valore solo nella colonna B
                        start_row, start_col, _, _ = self.get_merged_cell_range(self.foglio, i, 2)
                        self.foglio.cell(row=start_row, column=start_col, value=riga)
                    else:
                        self.foglio.cell(row=i, column=2, value=riga)  # Imposta la colonna B





    def crea_interfaccia(self):
        self.label_luogo_data = tk.Label(self.root, text="Luogo e Data:")
        self.label_luogo_data.grid(row=9, column=0, padx=10, pady=10)
        self.entry_luogo_data = tk.Entry(self.root, width=50)
        self.entry_luogo_data.grid(row=9, column=1, padx=10, pady=10)
        self.entry_luogo_data.insert(0, "Morlupo, lì ")

        self.label_informazione2 = tk.Label(self.root, text="Numero Doc. e luogo dei lavori: ")
        self.label_informazione2.grid(row=1, column=0, padx=10, pady=10)
        self.entry_informazione2 = tk.Entry(self.root, width=50)
        self.entry_informazione2.grid(row=1, column=1, padx=10, pady=10)
        self.entry_informazione2.insert(0, "Doc. N° — Prev. per —.")

        self.label_email = tk.Label(self.root, text="Email del committente:")
        self.label_email.grid(row=5, column=0, padx=10, pady=10)
        self.entry_email = tk.Entry(self.root, width=50)
        self.entry_email.grid(row=5, column=1, padx=10, pady=10)
        self.entry_email.insert(0, "email: ")

        self.label_residenza = tk.Label(self.root, text="Residenza del committente:")
        self.label_residenza.grid(row=3, column=0, padx=10, pady=10)
        self.combo_residenza = ttk.Combobox(self.root, values=["Via ", "P.zza ", "Largo ", ""], width=50)
        self.combo_residenza.grid(row=3, column=1, padx=10, pady=10)
        self.combo_residenza.set("Via ")
        
        self.label_cap = tk.Label(self.root, text="CAP, città e provincia:")
        self.label_cap.grid(row=4, column=0, padx=10, pady=10)
        self.combo_cap = ttk.Combobox(self.root, width=50)
        self.combo_cap.grid(row=4, column=1, padx=10, pady=10)
        self.combo_cap.set("")

        self.label_attenzione = tk.Label(self.root, text="Alla cortese attenzione:")
        self.label_attenzione.grid(row=2, column=0, padx=10, pady=10)
        self.combo_attenzione = ttk.Combobox(self.root, values=["Sig. ", "Sig.ra ", "Spett.le ", ""], width=50)
        self.combo_attenzione.grid(row=2, column=1, padx=10, pady=10)
        self.combo_attenzione.set("Sig. ")

        self.label_copia = tk.Label(self.root, text="Copia:")
        self.label_copia.grid(row=6, column=0, padx=10, pady=10)
        self.combo_copia = ttk.Combobox(self.root, values=["Copia installatore", "Copia committente", "Copia installatore e committente", ""], width=50)
        self.combo_copia.grid(row=6, column=1, padx=10, pady=10)
        self.combo_copia.set("Copia installatore")

        self.label_oggetto = tk.Label(self.root, text="Oggetto:")
        self.label_oggetto.grid(row=7, column=0, padx=10, pady=10)
        self.text_oggetto = tk.Text(self.root, height=5, width=50)
        self.text_oggetto.grid(row=7, column=1, padx=10, pady=10)

        # Aggiungi un contatore di caratteri
        self.label_contatore1 = tk.Label(self.root, text="Max 3 righe in oggetto. 80 caratteri a riga.")
        self.label_contatore1.grid(row=8, column=0, padx=10, pady=10)

        
        self.label_contatore = tk.Label(self.root, text="Caratteri rimanenti: 80")
        self.label_contatore.grid(row=8, column=1, padx=10, pady=10)

        # Altre etichette e campi di input ...





        #self.button_carica = tk.Button(self.root, text="Carica File Excel", command=self.carica_file_excel)
        #self.button_carica.grid(row=0, column=0, padx=10, pady=10)

        #self.button_modifica = tk.Button(self.root, text="Salva Excel", command=self.modifica_excel)
        #self.button_modifica.grid(row=10, column=1, padx=10, pady=10)





        self.label_risultato = tk.Label(self.root, text="")
        self.label_risultato.grid(row=11, column=0, columnspan=2, padx=10, pady=10)

        # Collega la funzione di conteggio dei caratteri all'evento di modifica del campo di testo
        self.text_oggetto.bind("<KeyRelease>", self.conta_caratteri)

    def conta_caratteri(self, event):
        testo = self.text_oggetto.get("1.0", "end-1c")
        lunghezza = len(testo)
        caratteri_rimasti = 80 - lunghezza
        self.label_contatore.config(text=f"Caratteri rimanenti: {caratteri_rimasti}")



    def get_merged_cell_range(self, foglio, row, column):
        for range_ in foglio.merged_cells.ranges:
            if range_.min_row <= row <= range_.max_row and range_.min_col <= column <= range_.max_col:
                return range_.min_row, range_.min_col, range_.max_row, range_.max_col
        return row, column, row, column
    def column_letter_to_index(self, letter):
        """Converte una lettera di colonna in un indice di colonna basato su zero."""
        return ord(letter.upper()) - ord("A")   
        
    def modifica_excel(self):
        luogo_data = self.entry_luogo_data.get()
        informazione2 = self.entry_informazione2.get()
        email = self.entry_email.get()
        residenza = self.combo_residenza.get()
        cap = self.combo_cap.get()
        attenzione = self.combo_attenzione.get()
        copia = self.combo_copia.get()
        oggetto = self.text_oggetto.get("1.0", "end-1c")  # Ottieni il testo da tutte le righe nel widget Text

        # Suddividi il testo dell'oggetto in righe
        righe_oggetto = oggetto.split('\n')

        if self.foglio:
            self.compila_informazioni(luogo_data, informazione2, email, residenza, cap, attenzione, copia, oggetto)

            for i, riga in enumerate(righe_oggetto[:3], start=16):
                cella_b = self.foglio.cell(row=i, column=2)
                cella_c = self.foglio.cell(row=i, column=3)
                cella_d = self.foglio.cell(row=i, column=4)
                cella_e = self.foglio.cell(row=i, column=5)

                # Verifica se la cella B è unita
                if self.is_cell_merged(self.foglio, i, 2):
                    # Trova la cella in alto a sinistra dell'intervallo unito e imposta il valore solo nella colonna B
                    start_row, start_col, _, _ = self.get_merged_cell_range(self.foglio, i, 2)
                    self.foglio.cell(row=start_row, column=start_col, value=riga)
                else:
                    # Imposta il valore nelle colonne B, C, D ed E
                    cella_b.value = riga
                    cella_c.value = riga
                    cella_d.value = riga
                    cella_e.value = riga

            self.foglio.parent.save("Preventivo Corrente.xlsx")
            self.label_risultato.config(text="File Excel modificato e salvato con successo!")



    def is_cell_merged(self, foglio, row, column):
        for merged_cell in foglio.merged_cells.ranges:
            if merged_cell.min_row <= row <= merged_cell.max_row and merged_cell.min_col <= column <= merged_cell.max_col:
                return True
        return False

    def unmerge_cell(self, foglio, row1, column1, row2=None, column2=None):
        if row2 is None:
            row2 = row1
        if column2 is None:
            column2 = column1

        # Verifica se la cella è unita prima di tentare di dividerla
        try:
            foglio.unmerge_cells(start_row=row1, start_column=column1, end_row=row2, end_column=column2)
        except KeyError:
            # La cella non è unita, quindi non è necessario dividerla
            pass

    def merge_cell(self, foglio, row1, column1, row2, column2):
        foglio.merge_cells(start_row=row1, start_column=column1, end_row=row2, end_column=column2)









def main():
    root = tk.Tk()
    app = ExcelEditor(root)
    root.mainloop()

if __name__ == "__main__":
    main()
