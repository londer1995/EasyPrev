import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl
import os
import tkinter.messagebox as messagebox  # Modifica questa linea
from openpyxl import load_workbook


class App:
    def crea_menu(self):
        self.menu = tk.Menu(self.root)
        self.root.config(menu=self.menu)

        self.filemenu = tk.Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="File", menu=self.filemenu)
        self.filemenu.add_command(label="Carica File Excel", command=self.carica_file_excel)
        self.filemenu.add_command(label="Salva Excel", command=self.salva_excel)  # Collega salva_excel al comando Salva Excel
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Informazioni", command=self.mostra_info)  # Nuovo comando per mostrare informazioni

    def mostra_info(self):
        messagebox.showinfo("EasyPrev AutoeXcel 1.0", " EasyPrev AutoeXcel è un software di preventivazione semi-automatizzata basata su python ed excel. Ultimo aggiornamnto 30.09.23.\n\n Programma realizzato per\n•Domenico Bertollini•\nSviluppato da\n•Luca Gualandri•")
        
    def __init__(self, root):
        self.root = root
        root.title("Interfaccia Materiali -  EasyPrev")
        self.canvas = None 

        # Imposta le dimensioni iniziali della finestra
        root.geometry("800x500")
        self.root.resizable(False, True)
        self.crea_menu()
        self.materials_data = {}
        self.materials_vars = []
        self.materials_widgets = []
        self.current_excel_row = None  # Inizializza la variabile di istanza, snnò non funzionaaaaaaa

        
        
        

        # Crea un frame per contenere tutti gli altri widget
        self.main_frame = tk.Frame(root)  # Imposta un colore di sfondo; sostituisci con un'immagine se preferisci
        root.grid_rowconfigure(0, weight=1)
        root.grid_columnconfigure(0, weight=1)
        self.main_frame.grid(sticky='nsew', padx=10, pady=10)  # Aggiunto un po' di padding
        self.main_frame.grid_rowconfigure(4, weight=1)  # Rende espandibile la riga 4
        self.main_frame.grid_columnconfigure(0, weight=1)  # Rende espandibile la colonna 0

        self.label_start_row = tk.Label(self.main_frame, 
                                        text="RIGA INIZIALE TABELLA MATERIALI - Controllare manualmente il preventivo in lavorazione ed inserire nella casella di seguito il numero della prima riga della tabella dei Materiali Occorrenti.",
                                        fg='red',  # colore del testo
                                        font='Arial 10 bold',  # font in grassetto
                                        wraplength=650)  # wrapping text
        self.label_start_row.grid(row=0, column=0, columnspan=5, sticky='w')

        self.start_row_var = tk.StringVar(value="CONTROLLARE_PREV._IN_LAVORAZIONE!")  # Valore predefinito impostato a 1
        vcmd = (root.register(self.validate), '%P')
        self.entry_start_row = tk.Entry(self.main_frame, textvariable=self.start_row_var, validate='key', validatecommand=vcmd)
        self.entry_start_row.grid(row=1, column=0, columnspan=5, sticky='w')
        self.confirm_button = tk.Button(self.main_frame, text="Conferma", command=self.confirm_start_row)
        self.confirm_button.grid(row=1, column=1, columnspan=5, pady=5)  # Posiziona il pulsante sopra il separatore

        self.separator = ttk.Separator(self.main_frame, orient='horizontal')
        self.separator.grid(row=2, column=0, columnspan=5, sticky='ew', pady=10)

        self.material_type = tk.StringVar()
        self.material_buttons = ttk.Combobox(self.main_frame, textvariable=self.material_type, values=["Quadro", "Impianto TV e Dati", "Automazioni"])
        
        
        self.material_buttons.grid(row=3, column=0, columnspan=5, sticky='w')
        self.material_buttons.bind("<<ComboboxSelected>>", self.update_materials_list)
        self.add_row_button = tk.Button(self.main_frame, text="Aggiungi riga", command=self.add_new_row)
        self.add_row_button.grid(row=3, column=4, sticky='w')  # sposta il pulsante vicino al menu a tendina
        self.remove_row_button = tk.Button(self.main_frame, text="Rimuovi ultima riga", command=self.remove_last_row)
        self.remove_row_button.grid(row=3, column=2, sticky='w')  # posiziona il pulsante vicino al pulsante "Aggiungi riga"



        self.canvas = tk.Canvas(self.main_frame)
        self.scrollbar = tk.Scrollbar(self.main_frame, orient="vertical", command=self.canvas.yview)
        self.canvas.grid(row=4, column=0, columnspan=5, sticky='nsew', pady=10)
        self.scrollbar.grid(row=4, column=5, sticky='nsew')  # Modifica 'ns' in 'nsew'
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)  # Scorrere con la rotellina del mouse
        self.canvas.bind_all('<MouseWheel>', self.on_mousewheel)  # Windows e MacOS
        self.canvas.bind_all('<Button-4>', self.on_mousewheel)  # Linux - rotellina su
        self.canvas.bind_all('<Button-5>', self.on_mousewheel)  # Linux - rotellina giù

        self.materials_frame = tk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.materials_frame, anchor="nw")


    def carica_file_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("File Excel", "*.xlsx")])
        if file_path:
            self.wb = load_workbook(file_path)  # Conserva l'intero workbook invece di solo il foglio attivo
            self.foglio = self.wb.active
            
    def load_data(self, material_type):
        material_to_filename = {
            "Quadro": "Quadro.xlsx",
            "Impianto_Di_Terra": "Impianto_Di_Terra.xlsx",
            "Automazioni": "Automazioni.xlsx"
        }
        filename = material_to_filename.get(material_type)
        if filename is None:
            print(f"Error: No file associated with {material_type}")
            return []

        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir, 'Materiali', filename)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        descriptions = [cell.value for row in sheet.iter_rows(min_col=1, max_col=1) for cell in row if cell.value]
        return descriptions
    
    def add_new_row(self):
        material_type = self.material_type.get()
        self.add_row(material_type)

        self.materials_frame.bind("<Configure>", lambda event: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.update_scrollregion()
        
    def remove_last_row(self):
        if self.materials_widgets:  # verifica che ci siano righe da rimuovere
            last_row_widgets = self.materials_widgets.pop()  # Rimuove gli widget dell'ultima riga dalla lista
            for widget in last_row_widgets:
                widget.destroy()  # Distrugge gli widget

            self.materials_vars.pop()  # Rimuove le variabili dell'ultima riga dalla lista

            # Riorganizza le righe
            self.rearrange_rows()

            # Aggiornare il canvas
            self.materials_frame.update_idletasks()
            self.canvas.config(scrollregion=self.canvas.bbox("all"))
            self.root.update()
            
            # Ri-configura la canvas per assicurarsi che lo scrollregion sia corretto
            self.on_configure(None)  # passare None poiché l'evento non è utilizzato
            self.update_scrollregion()

    def rearrange_rows(self):
        """Riorganizza i widget in self.materials_frame."""
        for index, widget_row in enumerate(self.materials_widgets):
            for col, widget in enumerate(widget_row):
                widget.grid(row=index + 1, column=col)

    def on_mousewheel(self, event):
        delta = event.delta if event.delta else -1 * event.num  # Differenziazione tra Windows/Mac e Linux
        self.canvas.yview_scroll(int(-1*(delta/120)), "units")  # Modifica per il funzionamento su Windows/Mac e Linux

    def on_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox('all'))
    
    def update_scrollregion(self):
        self.canvas.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))



    def confirm_start_row(self):
        start_row_value_str = self.start_row_var.get()
        if start_row_value_str.isdigit():  # controlla se la stringa è un numero intero
            start_row_value = int(start_row_value_str)
            self.current_excel_row = start_row_value  # Imposta la riga corrente del file Excel
            self.label_start_row.grid_forget()  # Nasconde il label
            self.entry_start_row.grid_forget()  # Nasconde l'entry
            self.confirm_button.grid_forget()   # Nasconde il button di conferma
        else:
            messagebox.showerror("Errore", "Inserisci un numero intero valido per la riga iniziale")

    def salva_excel(self):
        for materials_var in self.materials_vars:
            if materials_var['check'].get():
                if self.current_excel_row is None or self.foglio is None:
                    messagebox.showerror("Errore", "Selezionare una riga iniziale e un file Excel prima di salvare.")
                    return

        for materials_var in self.materials_vars:
            if materials_var['check'].get():
                self.foglio[f'B{self.current_excel_row}'] = materials_var['qty'].get()
                self.foglio[f'C{self.current_excel_row}'] = materials_var['desc'].get()
                self.foglio[f'D{self.current_excel_row}'] = materials_var['unit'].get()
                self.foglio[f'E{self.current_excel_row}'] = materials_var['total'].get()
                self.current_excel_row += 1  # Incrementa la riga corrente del file Excel

        # Ottieni il percorso della cartella principale del programma
        current_dir = os.path.dirname(os.path.abspath(__file__))

        # Crea il nome del file completo con il percorso
        file_path = os.path.join(current_dir, "Preventivo Corrente.xlsx")

        self.wb.save(file_path)  # Salva il file nella cartella principale
        messagebox.showinfo("Successo", "Dati salvati con successo nel file Excel.")


    def validate(self, value):
        if value == "":
            return True  # allow empty value
        try:
            int(value)
            return True
        except ValueError:
            return False

    def update_materials_list(self, event=None):
        self.save_current_data(self.material_type.get())  # Salvare i dati correnti prima di cambiare la lista

        for widget in self.materials_frame.winfo_children():
            widget.destroy()
        self.materials_vars.clear()

        material_type = self.material_type.get()
        self.material_type.set(material_type)  # Aggiorna la variabile del tipo di materiale

        descriptions = self.load_data(material_type)
        for description in descriptions:
            self.add_row(material_type, description)
        
        # Aggiunge una riga vuota alla fine per le nuove voci
        self.add_row(material_type)
        # Recupera i dati memorizzati per questa lista di materiali, se esistono
        materials_info = self.materials_data.get(material_type, None)
        if materials_info is None:
            # Se non ci sono dati memorizzati, carica i dati dal file
            descriptions = self.load_data(material_type)
            # Inizializza una nuova entry vuota alla fine della lista
            descriptions.append({"qty": "", "desc": "", "unit": "", "total": ""})
            self.materials_data[material_type] = descriptions
        else:
            descriptions = materials_info

    def update_row_state(self, var_check, combobox_qty, entry_desc, entry_unit, entry_total):
        state = "normal" if var_check.get() else "disabled"
        bg_color = "white" if var_check.get() else "grey"
        combobox_qty.config(state=state)
        entry_desc.config(state=state, bg=bg_color)
        entry_unit.config(state=state)
        entry_total.config(state=state)

    def update_total(self, combobox_qty, var_unit, var_total):
        try:
            qty = int(combobox_qty.get())
            unit = float(var_unit.get())
            total = qty * unit
            var_total.set(f'{total:.2f}')
        except ValueError:
            messagebox.showerror("Errore", "Inserisci valori numerici validi per la quantità e l'unità.")
            var_total.set('')
            
    def add_row(self, material_type, description=''):
        i = len(self.materials_vars) + 1
        var_check = tk.BooleanVar(value=False)  # Modificato a False
        var_desc = tk.StringVar(value=description)  # Imposta il valore della descrizione
        var_unit = tk.StringVar()
        var_total = tk.StringVar()

        combobox_qty = ttk.Combobox(self.materials_frame, values=[str(i) for i in range(1, 1000)], width=5)
        combobox_qty.set("1")
        entry_desc = tk.Entry(self.materials_frame, textvariable=var_desc, width=80)
        entry_unit = tk.Entry(self.materials_frame, textvariable=var_unit, width=15)
        entry_total = tk.Entry(self.materials_frame, textvariable=var_total, width=15)

        checkbutton = tk.Checkbutton(self.materials_frame, variable=var_check, 
            command=lambda vars=(var_check, combobox_qty, entry_desc, entry_unit, entry_total): 
            self.update_row_state(*vars))

        combobox_qty.bind("<<ComboboxSelected>>", lambda event, vars=(combobox_qty, var_unit, var_total): 
            self.update_total(*vars))
        var_unit.trace_add("write", lambda name, index, mode, vars=(combobox_qty, var_unit, var_total): 
            self.update_total(*vars))

        checkbutton.grid(row=i+1, column=0)
        combobox_qty.grid(row=i+1, column=1)
        entry_desc.grid(row=i+1, column=2)
        entry_unit.grid(row=i+1, column=3)
        entry_total.grid(row=i+1, column=4)

        #self.materials_vars.append((var_check, combobox_qty, var_desc, var_unit, var_total))
        self.materials_vars.append({
            'check': var_check,
            'desc': var_desc,
            'qty': combobox_qty,
            'unit': var_unit,
            'total': var_total
        })
        self.update_row_state(var_check, combobox_qty, entry_desc, entry_unit, entry_total)  # Aggiunto per impostare lo stato iniziale
        self.update_scrollregion()

        
    def save_current_data(self, material_type):  # Aggiungi material_type come argomento
        materials_info = []
        for materials_var in self.materials_vars:
            if materials_var['check'].get():
                material_info = {
                    'qty': vars['qty'].get(),
                    'desc': vars['desc'].get(),
                    'unit': vars['unit'].get(),
                    'total': vars['total'].get(),
                }
                materials_info.append(material_info)
        
        # Adesso, materials_info contiene tutte le informazioni. Potresti salvarlo in un file o in un database.
        # Esempio: salvataggio in un file JSON:
        import json
        with open(f'{material_type}_data.json', 'w') as f:
            json.dump(materials_info, f, indent=4)
        
    

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()